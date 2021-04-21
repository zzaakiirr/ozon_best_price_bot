import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


BEST_PRICE_STR = 'Есть дешевле, '
HEADERS = {
  'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'
}

# Excel file constants
EXCEL_PATH = 'products.xlsx'

URLS_COL_NAME = 'B'
CURRENT_PRICE_COL_NAME = 'C'
BEST_PRICE_COL_NAME = 'D'

START_INDEX = 2
END_INDEX = 1000000


def populate_sheet_with_product_prices(sheet,
                                       urls_col=URLS_COL_NAME,
                                       current_price_col=CURRENT_PRICE_COL_NAME,
                                       best_price_col=BEST_PRICE_COL_NAME,
                                       start_index=START_INDEX,
                                       end_index=END_INDEX):
  current_index = start_index

  for cell in sheet[f'{urls_col}{start_index}:{urls_col}{end_index}']:
    product_url = cell[0].value

    if not product_url:
      break

    soup = parse_html(product_url)
    current_price = find_current_price(soup)
    best_price = find_best_price(soup)

    print(product_url)
    print(f'\tCurrent price: {current_price}, Best price: {best_price}')

    sheet[f'{current_price_col}{current_index}'] = current_price
    sheet[f'{best_price_col}{current_index}'] = best_price

    current_index += 1


def get_sheet(workbook):
  sheet = workbook[workbook.sheetnames[0]] if len(workbook.sheetnames) else None

  if not sheet:
    print(f'Cannot get sheet from workbook in path {path}')
    return None

  return sheet


def parse_html(product_url, headers=HEADERS):
  response = requests.get(product_url, headers=headers)
  status_code = response.status_code

  if status_code != 200:
    print(f'Error while parsing product url: {product_url}, status_code: {status_code}')

  return BeautifulSoup(response.text, 'lxml')


def find_current_price(soup):
  target_tags = soup.select('.c2h5 span:first-child')

  if len(target_tags):
    return target_tags[0].contents[0] 

  print(f'Cannot parse current price')
  return None


def find_best_price(soup, target_str=BEST_PRICE_STR):
  target_divs = soup.find_all('div', class_='kxa6')

  for target_div in target_divs:
    div_text = target_div.contents[0] if len(target_div.contents) else ''
    if target_str in div_text:
      _, best_price = div_text.split(', ')
      return best_price

  return None


if __name__ == '__main__':
  workbook = load_workbook(EXCEL_PATH)
  sheet = get_sheet(workbook)
  populate_sheet_with_product_prices(sheet)
  workbook.save(EXCEL_PATH)
